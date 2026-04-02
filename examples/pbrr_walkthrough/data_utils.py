"""Data loading and preprocessing for PbrR walkthrough.

Provides shared utilities across all three steps: data loading from CSV,
tokenization, train/test splits, Pareto front identification, and
supplementary data loading (all-rounds dataset from Nature paper).
"""

import csv
from pathlib import Path
from typing import TypedDict

import numpy as np
import torch
from esm.tokenization.sequence_tokenizer import EsmSequenceTokenizer

DATA_FOLDER = Path("/data/ishan/pbbr/")
FC_CSV = DATA_FOLDER / "PbrR_Pb_Zn_FC.csv"
SUPPLEMENTARY_XLSX = DATA_FOLDER / "supplementary_data_1.xlsx"
SUPP_SOURCE_DATA = DATA_FOLDER / "supp_data_6.xlsx"
SAVED_SPLITS = DATA_FOLDER / "splits_dict.pt"
OUTPUTS_FOLDER = Path(__file__).resolve().parent / "outputs"
CHECKPOINTS_FOLDER = Path(__file__).resolve().parent / "checkpoints"

AMINO_ACIDS = "ACDEFGHIKLMNPQRSTVWY"
WT_SEQUENCE = "MNIQIGELAKRTACPVVTIRFYEQEGLLPPPGRSRGNFRLYGEEHVERLQFIRHCRSLDMPLSDVRTLLSYRKRPDQDCGEVNMLLDEHIRQVESRIGALLELKHHLVELREACSGARPAQSCGILQGLSDCVCDTRGTTAHPSD"


class PbrRData(TypedDict):
    """Round 1 data from the original guidance experiment."""

    mutations: list[list[tuple[int, str, str]]]  # [(pos, from_aa, to_aa), ...]
    pb_fc: np.ndarray  # (N,) Pb fold change
    zn_fc: np.ndarray  # (N,) Zn fold change
    sequences: list[str]  # protein sequences
    tokenized: torch.LongTensor  # (N, L+2) tokenized sequences with BOS/EOS
    wt_sequence: str


class SplitIndices(TypedDict):
    """Train/test split indices matching the original paper."""

    train: np.ndarray
    test: np.ndarray
    train_success: np.ndarray  # Pareto front subset of training data
    hard: list[int]  # multi-mutations at excluded positions


class AllRoundsData(TypedDict):
    """All rounds data from supplementary source data file."""

    mutations: list[list[tuple[int, str, str]]]
    pb_fc: np.ndarray
    zn_fc: np.ndarray
    sequences: list[str]
    tokenized: torch.LongTensor


def parse_mutation_string(mut_str: str) -> list[tuple[int, str, str]]:
    """Parse a mutation string like 'D64K_N83I' into [(64, 'D', 'K'), (83, 'N', 'I')]."""
    if mut_str == "WT" or mut_str == "Wildtype":
        return []
    muts = mut_str.split("_")
    return [(int(m[1:-1]), m[0], m[-1]) for m in muts]


def apply_mutations(wt_seq: str, muts: list[tuple[int, str, str]]) -> str:
    """Apply mutations to a wild-type sequence (1-indexed positions)."""
    seq_list = list(wt_seq)
    for pos, from_aa, to_aa in muts:
        assert seq_list[pos - 1] == from_aa, (
            f"Expected {from_aa} at position {pos}, got {seq_list[pos - 1]}"
        )
        seq_list[pos - 1] = to_aa
    return "".join(seq_list)


def load_round1_data() -> PbrRData:
    """Load round 1 data from the original PbrR_Pb_Zn_FC.csv."""
    tokenizer = EsmSequenceTokenizer()
    mutations, pb_fc, zn_fc, sequences = [], [], [], []

    with FC_CSV.open("r") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            mutations.append(parse_mutation_string(row[0]))
            pb_fc.append(float(row[1]))
            zn_fc.append(float(row[2]))
            sequences.append(row[3])

    pb_fc = np.array(pb_fc)
    zn_fc = np.array(zn_fc)

    tokenized = torch.cat(
        [tokenizer(seq, return_tensors="pt")["input_ids"] for seq in sequences], dim=0
    )

    return PbrRData(
        mutations=mutations,
        pb_fc=pb_fc,
        zn_fc=zn_fc,
        sequences=sequences,
        tokenized=tokenized,
        wt_sequence=sequences[-1],  # WT is the last entry
    )


def load_splits() -> SplitIndices:
    """Load the saved train/test splits from the original paper's experiment.

    The splits were computed by the original prepare_data_and_train_predictors.py
    script and saved to splits_dict.pt. Using the saved splits ensures exact
    reproducibility with the reference experiment.

    Split definitions:
    - train: all data except random half of multi-mutants (1060 entries)
    - val (test): random half of multi-mutants (39 entries)
    - train_success: training variants on the Pareto front (33 entries)
    - hard: multi-mutants at positions 64/104 with ≥3 mutations (15 entries)
    """
    import pickle as pkl

    d = pkl.load(SAVED_SPLITS.open("rb"))
    return SplitIndices(
        train=d["train"],
        test=d["val"],
        train_success=d["train_success"],
        hard=d["hard"],
    )


def find_ssm_positions(data: PbrRData) -> list[int]:
    """Find positions that were mutated to more than just alanine (SSM positions).

    Returns 1-indexed positions, matching the original experiment's mask positions.
    """
    mutations = data["mutations"]
    L = len(data["sequences"][0])

    only_ala_positions = [
        p
        for p in range(1, L + 1)
        if {
            AMINO_ACIDS.index(aa)
            for muts in mutations
            for (pos, _, aa) in muts
            if pos == p
        }
        == {AMINO_ACIDS.index("A")}
        or not any(pos == p for muts in mutations for (pos, _, _) in muts)
    ]
    return [i for i in range(1, L + 1) if i not in only_ala_positions]


def make_masked_wt(data: PbrRData, ssm_positions: list[int]) -> torch.LongTensor:
    """Create a masked WT sequence for generation (mask at SSM positions).

    Returns tokenized sequence (L+2,) with BOS/EOS and mask tokens at SSM positions.
    SSM positions are 1-indexed protein positions; token positions are offset by +1
    for the BOS token.
    """
    tokenizer = EsmSequenceTokenizer()
    wt_tokens = tokenizer(data["wt_sequence"], return_tensors="pt")[
        "input_ids"
    ].squeeze(0)
    for pos in ssm_positions:
        wt_tokens[pos] = (
            tokenizer.mask_token_id
        )  # +0 because ESM tokenizer already adds BOS
    return wt_tokens


def load_all_rounds_data() -> AllRoundsData:
    """Load all-rounds FC data from the supplementary source data file.

    Extracts round 1 FC data from Figure 3 Panel E and later rounds from
    Figure 4 panels. The WT sequence is used as the base for applying mutations.
    """
    import openpyxl

    tokenizer = EsmSequenceTokenizer()
    wb = openpyxl.load_workbook(str(SUPP_SOURCE_DATA))

    all_fc: dict[str, tuple[float, float]] = {}

    # Round 1: Figure 3 Panel E (columns 29-31)
    ws3 = wb["Figure 3"]
    for row in range(3, ws3.max_row + 1):
        mutant = ws3.cell(row=row, column=29).value
        pb = ws3.cell(row=row, column=30).value
        zn = ws3.cell(row=row, column=31).value
        if mutant is not None and pb is not None:
            all_fc[mutant] = (float(pb), float(zn))

    # Later rounds: Figure 4 panels B, C (cols 10-12), D (18-20), E (22-24), F (30-32)
    ws4 = wb["Figure 4"]
    panel_cols = [(1, 2, 3), (10, 11, 12), (18, 19, 20), (22, 23, 24), (30, 31, 32)]
    for mut_col, pb_col, zn_col in panel_cols:
        for row in range(4, ws4.max_row + 1):
            mutant = ws4.cell(row=row, column=mut_col).value
            pb = ws4.cell(row=row, column=pb_col).value
            zn = ws4.cell(row=row, column=zn_col).value
            if mutant is not None and pb is not None and zn is not None:
                try:
                    all_fc[mutant] = (float(pb), float(zn))
                except (ValueError, TypeError):
                    continue  # skip non-numeric entries (headers, etc.)

    # Build data arrays, applying mutations to WT
    mutations, pb_fc, zn_fc, sequences = [], [], [], []
    skipped = 0
    for mut_str, (pb, zn) in all_fc.items():
        try:
            muts = parse_mutation_string(mut_str)
            seq = apply_mutations(WT_SEQUENCE, muts)
            mutations.append(muts)
            pb_fc.append(pb)
            zn_fc.append(zn)
            sequences.append(seq)
        except (AssertionError, ValueError, IndexError, KeyError):
            skipped += 1
            continue

    if skipped > 0:
        print(f"Skipped {skipped} entries due to mutation parsing errors")

    tokenized = torch.cat(
        [tokenizer(seq, return_tensors="pt")["input_ids"] for seq in sequences], dim=0
    )

    return AllRoundsData(
        mutations=mutations,
        pb_fc=np.array(pb_fc),
        zn_fc=np.array(zn_fc),
        sequences=sequences,
        tokenized=tokenized,
    )


def ohe_from_tokenized(
    tokenized: torch.LongTensor,
    wt_tokenized: torch.LongTensor | None = None,
) -> torch.FloatTensor:
    """Convert tokenized sequences to one-hot encoding, optionally centered on WT.

    Args:
        tokenized: (N, L+2) tokenized sequences
        wt_tokenized: If provided, center OHE around WT (X_ohe - WT_ohe)

    Returns:
        (N, (L+2)*V) flattened OHE features
    """
    tokenizer = EsmSequenceTokenizer()
    V = tokenizer.vocab_size
    X_ohe = torch.nn.functional.one_hot(tokenized, num_classes=V).float()
    X_flat = X_ohe.view(X_ohe.size(0), -1)

    if wt_tokenized is not None:
        WT_ohe = torch.nn.functional.one_hot(wt_tokenized, num_classes=V).float()
        WT_flat = WT_ohe.view(1, -1)
        X_flat = X_flat - WT_flat

    return X_flat
